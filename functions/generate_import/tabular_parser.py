#!/usr/bin/env python3
"""
Deterministic, never-raising tabular file parser — the hardening layer
described in task-parser-hardening.txt. Sniffs actual file content (never
trusts the extension), iterates every sheet, scores candidate header rows
against normalize.py's alias dictionary, and normalizes every cell — without
ever calling out to an LLM and without ever raising to the caller.

Core contract: parse_tabular_file() ALWAYS returns a ParseResult dict:
    {rows, warnings, errors, file_type_detected, sheets_parsed, sheets_skipped}
0 rows is always accompanied by a non-empty `errors` entry explaining why.
Every stage (file -> sheets -> header -> rows -> cells) is wrapped in its own
try/except so a failure at any granularity degrades to a reported error at
that granularity instead of taking down anything above it.

Currently wired into main.py for the `students` entity only — that's the one
whose column alias dictionary (normalize.STUDENT_HEADER_ALIASES) this module
was built against. The engine itself (sniffing/sheets/header-scoring) is
schema-agnostic and takes schema_keys/field_aliases/required_field as
parameters, so a future tabular teacher/subject import can reuse it without
copying any of this logic — only normalize.py's alias dict and normalizers
are required to be shared per the task's NOTES section.
"""
import csv
import io
import re
from datetime import date, datetime
from html.parser import HTMLParser

from normalize import (
    canonicalize, clean_email, clean_gender, clean_name, clean_phone,
    collapse_ws, extract_grade_tokens, normalize_section, parse_dob_flexible,
    build_alias_lookup, fuzzy_best_match,
)


class UnreadableFileError(Exception):
    """Raised internally by the loaders; always caught and turned into a
    ParseResult-level error — never escapes parse_tabular_file()."""


# ------------------------------------------------------------- sniffing ----
def sniff_file_type(raw):
    """Content-sniffed file type, ignoring any extension entirely. Returns
    one of: 'empty', 'xlsx', 'ole' (legacy xls OR encrypted xlsx — disambig-
    uated by the loader), 'html', 'text' (csv/tsv candidate)."""
    if not raw:
        return "empty"
    if raw[:4] in (b"PK\x03\x04", b"PK\x05\x06", b"PK\x07\x08"):
        return "xlsx"
    if raw[:8] == b"\xD0\xCF\x11\xE0\xA1\xB1\x1A\xE1":
        return "ole"
    sample = raw[:4096].decode("utf-8", errors="replace").lstrip("﻿ \t\r\n")
    if re.match(r"(?is)^\s*(<!doctype\s+html|<html|<table)", sample):
        return "html"
    return "text"


def decode_bytes(raw):
    """utf-8 -> utf-8-sig -> cp1252 -> latin-1 chain; latin-1 accepts every
    byte so this never raises (errors='replace' is unreachable in practice
    but kept as a documented last resort, per the task contract)."""
    for enc in ("utf-8", "utf-8-sig", "cp1252", "latin-1"):
        try:
            return raw.decode(enc), enc
        except UnicodeDecodeError:
            continue
    return raw.decode("latin-1", errors="replace"), "latin-1(replace)"


def detect_delimiter(text):
    sample = "\n".join(text.splitlines()[:10])
    try:
        return csv.Sniffer().sniff(sample, delimiters=",;\t|").delimiter
    except Exception:
        counts = {d: sample.count(d) for d in (",", "\t", ";")}
        best = max(counts, key=counts.get)
        return best if counts[best] > 0 else ","


# --------------------------------------------------------------- loaders ---
def _cell_to_str(v):
    if v is None:
        return ""
    if isinstance(v, float):
        return str(int(v)) if v.is_integer() else str(v)
    return str(v).strip()


def _load_xlsx(raw):
    import openpyxl
    try:
        wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    except Exception as e:
        raise UnreadableFileError(f"could not open as xlsx: {e}")
    sheets, warnings = [], []
    for name in wb.sheetnames:
        try:
            ws = wb[name]
            rows = [list(r) for r in ws.iter_rows(values_only=True)]
            hidden = getattr(ws, "sheet_state", "visible") != "visible"
        except Exception as e:
            warnings.append({"sheet": name, "row": None, "field": None,
                              "message": f"could not read sheet: {e}", "severity": "warning"})
            continue
        sheets.append((name, rows, hidden))
    return sheets, warnings


def _xlrd_cell_value(cell, datemode):
    import xlrd
    if cell.ctype == xlrd.XL_CELL_DATE:
        try:
            return xlrd.xldate_as_datetime(cell.value, datemode)
        except Exception:
            return cell.value
    return cell.value


def _load_ole(raw):
    """D0CF11E0 compound-file signature covers both legacy .xls and
    password-encrypted .xlsx (OOXML wrapped in an OLE container) — try xlrd
    first, then disambiguate the failure."""
    import xlrd
    try:
        book = xlrd.open_workbook(file_contents=raw)
    except Exception as e:
        if "EncryptedPackage".encode("utf-16-le") in raw or "password" in str(e).lower() or "encrypt" in str(e).lower():
            raise UnreadableFileError("password-protected file")
        raise UnreadableFileError(f"corrupt or unsupported legacy .xls file: {e}")

    sheets, warnings = [], []
    for sheet in book.sheets():
        try:
            rows = [[_xlrd_cell_value(sheet.cell(r, c), book.datemode) for c in range(sheet.ncols)]
                    for r in range(sheet.nrows)]
        except Exception as e:
            warnings.append({"sheet": sheet.name, "row": None, "field": None,
                              "message": f"could not read sheet: {e}", "severity": "warning"})
            continue
        hidden = getattr(sheet, "visibility", 0) != 0
        sheets.append((sheet.name, rows, hidden))
    return sheets, warnings


class _HTMLTableParser(HTMLParser):
    """Minimal <table> extractor — stdlib only, no lxml/pandas dependency.
    Handles colspan (repeats the cell text); rowspan is NOT carried down into
    following rows (acceptable gap — real-world school HTML exports almost
    never rowspan a header, and a misparsed extra sheet just gets skipped
    with a reason downstream, never crashes)."""

    def __init__(self):
        super().__init__()
        self.tables = []
        self._depth = 0
        self._table = None
        self._row = None
        self._cell = None
        self._cell_attrs = None

    def handle_starttag(self, tag, attrs):
        if tag == "table":
            self._depth += 1
            if self._depth == 1:
                self._table = []
        elif tag == "tr" and self._table is not None:
            self._row = []
        elif tag in ("td", "th") and self._row is not None:
            self._cell = []
            self._cell_attrs = dict(attrs)

    def handle_endtag(self, tag):
        if tag in ("td", "th") and self._cell is not None:
            text = collapse_ws("".join(self._cell))
            try:
                colspan = max(1, int(self._cell_attrs.get("colspan", 1) or 1))
            except ValueError:
                colspan = 1
            self._row.extend([text] * colspan)
            self._cell = None
        elif tag == "tr" and self._row is not None:
            self._table.append(self._row)
            self._row = None
        elif tag == "table":
            self._depth = max(0, self._depth - 1)
            if self._depth == 0 and self._table is not None:
                self.tables.append(self._table)
                self._table = None

    def handle_data(self, data):
        if self._cell is not None:
            self._cell.append(data)


def _load_html(raw):
    text, enc = decode_bytes(raw)
    parser = _HTMLTableParser()
    try:
        parser.feed(text)
    except Exception as e:
        raise UnreadableFileError(f"malformed HTML: {e}")
    if not parser.tables:
        raise UnreadableFileError("no <table> found in HTML content")
    return [(f"Table{i + 1}", rows, False) for i, rows in enumerate(parser.tables)], []


def _load_text(raw):
    text, enc = decode_bytes(raw)
    warnings = []
    if enc not in ("utf-8", "utf-8-sig"):
        warnings.append({"sheet": None, "row": None, "field": None,
                          "message": f"decoded using fallback encoding '{enc}'", "severity": "warning"})
    if not text.strip():
        raise UnreadableFileError("file has no parseable text content")
    delim = detect_delimiter(text)
    rows = list(csv.reader(io.StringIO(text), delimiter=delim))
    resolved_type = "tsv" if delim == "\t" else "csv"
    return [("Sheet1", rows, False)], warnings, resolved_type


# ------------------------------------------------------------- ParseResult --
def _new_result(file_type):
    return {"rows": [], "warnings": [], "errors": [],
            "file_type_detected": file_type, "sheets_parsed": [], "sheets_skipped": [],
            "unmapped_headers": []}


def _add_error(result, message, sheet=None, row=None, field=None):
    result["errors"].append({"sheet": sheet, "row": row, "field": field,
                              "message": message, "severity": "error"})


def _best_zero_row_reason(result):
    if result["sheets_skipped"] and not result["sheets_parsed"]:
        reasons = sorted({s["reason"] for s in result["sheets_skipped"]})
        return "no usable sheet: " + "; ".join(reasons)
    if result["sheets_parsed"]:
        return "header found but no data rows in " + ", ".join(result["sheets_parsed"])
    return "no recognizable header in any sheet"


def _finalize(result):
    if not result["rows"] and not result["errors"]:
        _add_error(result, _best_zero_row_reason(result))
    return result


# ---------------------------------------------------------- header detect --
def _is_row_empty(row):
    return all(c is None or str(c).strip() == "" for c in row)


def _is_sheet_effectively_empty(rows):
    return not rows or all(_is_row_empty(r) for r in rows)


def _row_text_key(row):
    return "|".join(canonicalize(_cell_to_str(c)) for c in row)


def _fill_forward(row):
    """Forward-fills blank cells from the last non-blank one — recovers
    merged header cells where only the first column of a merged group holds
    text (openpyxl/xlrd both report the rest as None for a merge)."""
    out, last = [], ""
    for c in row:
        v = _cell_to_str(c)
        if v:
            last = v
        out.append(last)
    return out


def _join_header_rows(row1, row2):
    f1 = _fill_forward(row1)
    width = max(len(f1), len(row2))
    joined = []
    for i in range(width):
        a = f1[i] if i < len(f1) else ""
        b = _cell_to_str(row2[i]) if i < len(row2) else ""
        joined.append(f"{a} {b}".strip() if a and b else (a or b))
    return joined


_PARENTHETICAL_RE = re.compile(r"\([^)]*\)")


def _strip_parenthetical(canon):
    """'gr/emis/sts(uniqueno)' -> 'gr/emis/sts'.

    Source headers routinely carry a trailing qualifier the alias list has no
    reason to enumerate — '(unique No.)', '(if any)', '(DD/MM/YYYY)'. It is
    dead weight for matching but not for scoring: Hillgreen's
    'GR / EMIS / STS (unique No.)' fuzzy-matched its own alias at 0.688,
    under the 0.84 threshold, so a field the school explicitly asked to keep
    was dropped as an unmapped column.

    Returns `canon` unchanged when stripping would empty it — a header that is
    nothing but a parenthetical has no bare form to match on.
    """
    out = _PARENTHETICAL_RE.sub("", canon).strip()
    return out or canon


def _match_field(canon, alias_lookup):
    """Canonicalized header text -> schema field, or None.

    Exact first, then fuzzy, then both again without a trailing parenthetical.
    Ordered so the bare form is only ever a fallback: it can rescue a header
    the full form missed, never override one the full form already matched.
    """
    field = alias_lookup.get(canon)
    if field is not None:
        return field
    candidate, ratio = fuzzy_best_match(canon, {c: c for c in alias_lookup})
    if candidate and ratio >= 0.84:
        return alias_lookup[candidate]

    bare = _strip_parenthetical(canon)
    if bare == canon:
        return None
    field = alias_lookup.get(bare)
    if field is not None:
        return field
    candidate, ratio = fuzzy_best_match(bare, {c: c for c in alias_lookup})
    if candidate and ratio >= 0.84:
        return alias_lookup[candidate]
    return None


def _score_header_row(row, alias_lookup):
    col_map, labels, unmapped = {}, [], []
    used_fields = set()
    score = 0
    for idx, cell in enumerate(row):
        text = _cell_to_str(cell)
        labels.append(text)
        if not text:
            continue
        field = _match_field(canonicalize(text), alias_lookup)
        if field is None or field in used_fields:
            if text:
                unmapped.append(text)
            continue
        col_map[idx] = field
        used_fields.add(field)
        score += 1
    return score, col_map, labels, unmapped


def _detect_header(rows, alias_lookup, required_field, min_score, max_scan=20):
    n = min(len(rows), max_scan)
    best = None  # (score, start_idx, span, col_map, labels, unmapped)
    for i in range(n):
        score, col_map, labels, unmapped = _score_header_row(rows[i], alias_lookup)
        if required_field in col_map.values() and (best is None or score > best[0]):
            best = (score, i, 1, col_map, labels, unmapped)
        if i + 1 < n:
            joined = _join_header_rows(rows[i], rows[i + 1])
            score2, col_map2, labels2, unmapped2 = _score_header_row(joined, alias_lookup)
            if required_field in col_map2.values() and (best is None or score2 > best[0]):
                best = (score2, i, 2, col_map2, labels2, unmapped2)
    if best is None or best[0] < min_score:
        return None
    return best


_CLASS_HINT_RE = re.compile(r"(?i)^\s*(class|grade|std|section|sec)\s*[:\-]?\s*(.+)$")


def _extract_metadata_hints(rows_above_header):
    """Per-sheet metadata rows like 'Class: I B' above the header -> a
    grade/section hint applied to rows missing those values (common pattern:
    one sheet per section)."""
    hints = {}
    for row in rows_above_header:
        joined = " ".join(_cell_to_str(c) for c in row if _cell_to_str(c))
        m = _CLASS_HINT_RE.match(joined)
        if not m:
            continue
        label, rest = m.group(1).lower(), m.group(2).strip()
        if label in ("section", "sec"):
            hints["section"] = normalize_section(rest)
            continue
        tokens = extract_grade_tokens(rest)
        if tokens:
            hints["grade"] = tokens[0]
            remainder = rest
            for tok in (tokens[0],):
                remainder = re.sub(re.escape(tok), "", remainder, flags=re.IGNORECASE)
            sec_match = re.search(r"\b([A-Za-z])\b\s*$", remainder.strip())
            if sec_match:
                hints["section"] = normalize_section(sec_match.group(1))
    return hints


_SUMMARY_ROW_RE = re.compile(r"(?i)^(grand\s+total|total|sum|count)\b")


def _dedupe_labels(labels):
    seen, out = {}, []
    for label in labels:
        if not label:
            out.append(label)
            continue
        seen[label] = seen.get(label, 0) + 1
        out.append(label if seen[label] == 1 else f"{label}_{seen[label]}")
    return out


# ----------------------------------------------------------- cell normalize
def _normalize_cell(field, raw_val):
    if field in ("student_name", "father_name", "mother_name"):
        v, _fix = clean_name(_cell_to_str(raw_val))
        return v, None
    if field == "email":
        text = _cell_to_str(raw_val)
        v, _fix, valid = clean_email(text)
        return v, (f"invalid email: '{text}'" if text and not valid else None)
    if field == "dob":
        return parse_dob_flexible(raw_val)
    if field == "contact":
        return clean_phone(_cell_to_str(raw_val))
    if field == "gender":
        return clean_gender(_cell_to_str(raw_val))
    if field == "grade":
        text = _cell_to_str(raw_val)
        tokens = extract_grade_tokens(text)
        if tokens:
            warn = f"multiple grades found in '{text}' — using '{tokens[0]}'" if len(tokens) > 1 else None
            return tokens[0], warn
        return text, (f"unrecognized grade value: '{text}'" if text else None)
    if field == "section":
        return normalize_section(_cell_to_str(raw_val)), None
    return collapse_ws(_cell_to_str(raw_val)), None


# --------------------------------------------------------------- sheet/row -
def _parse_row(raw_row, col_map, metadata_hints, sheet_name, row_num, hidden,
               result, header_text_key, required_field, schema_keys):
    if _is_row_empty(raw_row):
        return
    if _row_text_key(raw_row) == header_text_key:
        return  # repeated header pasted mid-sheet — silently dropped, debug-level
    first_val = next((_cell_to_str(c) for c in raw_row if _cell_to_str(c)), "")
    if _SUMMARY_ROW_RE.match(first_val):
        return  # summary row ("Total: 45") — silently dropped, debug-level

    data = {k: "" for k in schema_keys}
    cell_warnings = []
    for idx, field in col_map.items():
        if field not in schema_keys:
            # Recognized (counts for header scoring / metadata hints) but
            # deliberately out-of-schema, e.g. "address" — see normalize.
            # STUDENT_HEADER_ALIASES. Never persisted, belt-and-braces same
            # as extract_file()'s BANNED_KEYS filter on the LLM path.
            continue
        raw_val = raw_row[idx] if idx < len(raw_row) else None
        try:
            value, warn = _normalize_cell(field, raw_val)
        except Exception as e:
            value, warn = "", f"could not parse {field}: {e}"
        data[field] = value
        if warn:
            cell_warnings.append((field, warn))

    for hint_field in ("grade", "section"):
        if hint_field in metadata_hints and not (data.get(hint_field) or "").strip():
            data[hint_field] = metadata_hints[hint_field]

    for field, warn in cell_warnings:
        result["warnings"].append({"sheet": sheet_name, "row": row_num, "field": field,
                                    "message": warn, "severity": "warning"})

    excluded = not (data.get(required_field) or "").strip()
    if excluded:
        # Kept (not dropped) so the row stays visible/auditable in review —
        # "excluded" per the task contract, never silently vanished — but
        # still reported as the one error condition that blocks commit.
        result["errors"].append({
            "sheet": sheet_name, "row": row_num, "field": required_field,
            "message": f"missing {required_field.replace('_', ' ')}", "severity": "error",
        })
    result["rows"].append({"data": data, "sheet": sheet_name, "row": row_num,
                            "hidden_sheet": hidden, "excluded": excluded})


def _parse_sheet(sheet_name, rows, hidden, result, alias_lookup, required_field,
                  min_header_score, schema_keys):
    if _is_sheet_effectively_empty(rows):
        return  # genuinely empty sheets are skipped silently, per contract

    best = _detect_header(rows, alias_lookup, required_field, min_header_score)
    if best is None:
        result["sheets_skipped"].append({"name": sheet_name, "reason": "no recognizable header"})
        return

    score, start_idx, span, col_map, labels, unmapped = best
    result["unmapped_headers"].extend(_dedupe_labels(unmapped))
    metadata_hints = _extract_metadata_hints(rows[:start_idx])
    header_text_key = _row_text_key(labels)
    result["sheets_parsed"].append(sheet_name)

    data_rows = rows[start_idx + span:]
    for offset, raw_row in enumerate(data_rows):
        row_num = start_idx + span + offset + 1  # 1-indexed source row number
        try:
            _parse_row(raw_row, col_map, metadata_hints, sheet_name, row_num, hidden,
                       result, header_text_key, required_field, schema_keys)
        except Exception as e:
            result["errors"].append({"sheet": sheet_name, "row": row_num, "field": None,
                                      "message": f"row failed to parse: {e}", "severity": "error"})


# ------------------------------------------------------------------- entry -
def parse_tabular_file(filename, raw, schema_keys, field_aliases, required_field,
                        min_header_score=2):
    """Never raises. Always returns a ParseResult dict — see module docstring.
    `schema_keys`: the full set of fields every output row's `data` dict will
    have (missing ones default to ''). `field_aliases`: {field: [alias
    phrases]} for header scoring (see normalize.STUDENT_HEADER_ALIASES).
    `required_field`: the one field whose absence excludes a row entirely."""
    try:
        ftype = sniff_file_type(raw)
    except Exception as e:
        result = _new_result("unknown")
        _add_error(result, f"file unreadable: {e}")
        return _finalize(result)

    if ftype == "empty":
        result = _new_result("empty")
        _add_error(result, "empty file")
        return _finalize(result)

    try:
        if ftype == "xlsx":
            sheets, load_warnings = _load_xlsx(raw)
        elif ftype == "ole":
            sheets, load_warnings = _load_ole(raw)
            ftype = "xls"
        elif ftype == "html":
            sheets, load_warnings = _load_html(raw)
        else:
            sheets, load_warnings, resolved = _load_text(raw)
            ftype = resolved
    except UnreadableFileError as e:
        result = _new_result(ftype)
        _add_error(result, f"file unreadable: {e}")
        return _finalize(result)
    except Exception as e:
        result = _new_result(ftype)
        _add_error(result, f"file unreadable: {e}")
        return _finalize(result)

    result = _new_result(ftype)
    result["warnings"].extend(load_warnings)

    if not sheets:
        _add_error(result, "no sheets found in file")
        return _finalize(result)

    alias_lookup = build_alias_lookup(field_aliases)
    for sheet_name, sheet_rows, hidden in sheets:
        try:
            _parse_sheet(sheet_name, sheet_rows, hidden, result, alias_lookup,
                         required_field, min_header_score, schema_keys)
        except Exception as e:
            result["sheets_skipped"].append({"name": sheet_name, "reason": f"sheet parse failed: {e}"})

    return _finalize(result)
